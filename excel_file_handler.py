from openpyxl import Workbook, load_workbook
import time
from fpdf import FPDF
import sys


class Excel_File_Handler:

    def __init__(self):
        """
            This is Excel_File_Handler Constructor
            this method create three excel files for
                1. Customer
                2. food
                3. Room
            also adds header data on those files 
        """

        self.customer_workbook = Workbook()
        self.room_workbook = Workbook()
        self.food_workbook = Workbook()

        self.customer_worksheet = self.customer_workbook.active
        self.room_sheet = self.room_workbook.active
        self.food_sheet = self.food_workbook.active

        self.customer_header_data = [
            'ID', 'Name', 'Address', 'Check-In-Date', 'Check-Out-Date', 'Room-ID', 'Food-ID']
        self.room_header_data = ['ID', 'Room-Number', 'Price', 'Is_Reserved']
        self.food_header_data = ['ID', 'Name', 'Price']

        self.customer_worksheet.append(self.customer_header_data)
        self.room_sheet.append(self.room_header_data)
        self.food_sheet.append(self.food_header_data)

        self.customer_workbook.save('customer_excel_file.xlsx')
        self.room_workbook.save('room_excel_file.xlsx')
        self.food_workbook.save('food_excel_file.xlsx')

    def load_customer_workbook(self):
        """
            This method loads the Customer excel file 
            it is not creating but load to reuse the this part of code in multiple places
        """

        self.customer_book = load_workbook('customer_excel_file.xlsx')
        self.customer_sheet = self.customer_book.active

    def load_room_workbook(self):
        """
            This method loads the room excel file 
            it is not creating but load to reuse the this part of code in multiple places
        """

        self.room_book = load_workbook('room_excel_file.xlsx')
        self.room_sheet = self.room_book.active

    def load_food_workbook(self):
        """
            This method loads the food excel file 
            it is not creating but load to reuse the this part of code in multiple places
        """

        self.food_book = load_workbook('food_excel_file.xlsx')
        self.food_sheet = self.food_book.active

    def customer_inputs(self):
        """
            This method takes Customer data from the command prompt
            which is:
                1. ID of Customer
                2. Name of Customer
                3. Address of Customer
            also we reused this method in different methods
        """

        self.cust_id = input('Enter Customer ID: ')
        self.cust_name = input('Enter Customer Name: ')
        self.cust_address = input('Enter Customer Address: ')

    def room_inputs(self):
        """
            This method takes Room data from the command prompt
            which is:
                1. ID of Room
                2. Name of Room
                3. Price of Food
            also we reused this method in different methods
        """

        self.room_id = input('Enter Room ID: ')
        self.room_number = input('Enter Room Number: ')
        self.room_price = input('Enter Room Price: ')

    def food_inputs(self):
        """
            This method takes Food data from the command prompt
            which is:
                1. ID of Food
                2. Room no of Food
                3. Price of Food
            also we reused this method in different methods
        """
        self.food_id = input('Enter Food ID: ')
        self.food_name = input('Enter Food Name: ')
        self.food_price = input('Enter Food Price: ')

    def add_customer(self):
        """
            This method add Customer data in Excel files

            then checks whether there is room if so. then list the rooms
            and user can assign for the customer otherwise enable the user
            to add manually room data and assign for the current customer

            and finally checks whether there is food availabe if so. then list all the foods
            and user can assign the foods based on ID for the customer
            otherwise this method enable the user to add manually food and the food will assign
            for the customer.              
        """

        self.load_customer_workbook()
        self.customer_inputs()

        self.r_id = ''
        self.f_id = ''

        if self.check_if_room_exists():
            self.list_rooms()
            self.r_id = input('Enter the Room ID for the Customer: ')

        else:
            print('\nNot Available Rooms\n')
            self.r_id = self.add_room()

        if self.check_if_food_exists():

            print()
            self.list_foods()
            print()

            self.f_id = input('Enter the Food ID for the Customer: ')

        else:

            self.f_id = self.add_food()

        self.customer_sheet.append(
            [
                self.cust_id, self.cust_name, self.cust_address,
                time.strftime('%Y-%m-%d %H:%m:%S'), '',
                self.r_id, self.f_id
            ])

        self.customer_book.save('customer_excel_file.xlsx')
        self.reserve_room(int(self.r_id) + 1)

    def list_customers(self):
        """
            list_customers method list or fetch all the customers from the customer_excel_file
            and pass the data to print_data method and the mentioned method will print the customers
        """
        
        self.load_customer_workbook()
        self.customer_sheet.dimensions
        self.print_data(self.customer_sheet.max_row,
                        self.customer_worksheet.max_column, self.customer_sheet)

    def add_room(self):
        """
            add_room method first obtain the room data from the prompt
            then insert those data in room_excel_file.xlsx
        """
        
        self.load_room_workbook()
        self.room_inputs()
        self.room_sheet.append(
            [self.room_id, self.room_number, self.room_price, 'No'])
        self.room_book.save('room_excel_file.xlsx')
        return self.room_id

    # Done.
    def list_rooms(self):
        """
            list_rooms method select all data from the room_excel_file.xlsx and
            then pass those data in print_data to show
        """

        self.load_room_workbook()
        self.room_sheet.dimensions
        self.print_data(self.room_sheet.max_row,
                        self.room_sheet.max_column, self.room_sheet)

    # For Reusing...

    def print_data(self, maximum_row, maximum_column, sheet):
        """ print_data method actually show data in terminal in a particular appearance

        Args:
            maximum_row (Integer): Maximum row of data in excel file
            maximum_column (Integer): Maximum column of data in excel file
            sheet (Sheet Object): this the object of the sheet which we are going to loop and get the data
        """
        print()
        for r in range(1, maximum_row + 1):
            for c in range(1, maximum_column + 1):
                print(sheet.cell(row=r, column=c).value, end='	  ')
            print()
        print()

    # Done.

    def add_food(self):
        """
            add_food method first obtain the food data from the prompt
            then insert those data in food_excel_file.xlsx
        """

        self.load_food_workbook()
        self.food_inputs()
        self.food_sheet.append([self.food_id, self.food_name, self.food_price])
        self.food_book.save('food_excel_file.xlsx')
        return self.food_id

    # Done.
    def list_foods(self):
        """
            list_foods method select all data from the food_excel_file.xlsx and
            then pass those data in print_data to show
        """
        
        self.load_food_workbook()
        self.food_sheet.dimensions
        print('\nAvailable Food Lists:\n')
        self.print_data(self.food_sheet.max_row,
                        self.food_sheet.max_column, self.food_sheet)

    def reserve_room(self, room_id):
        """ this method modifies the column D[*] based on id which is passed
        and modify and change the value of that specific column and row to (Yes) means theat room reserved

        Args:
            room_id (string): specify the column in excel file based on the number which is passed
        """
        
        self.load_room_workbook()
        self.room_sheet.dimensions
        self.room_sheet['D{id}'.format(id=str(room_id))] = 'Yes'
        self.room_book.save('room_excel_file.xlsx')

    # Done. (IF Room Exists Return TRUE, Else Return FALSE)

    def check_if_room_exists(self):
        """
            check_if_room_exists method checks whether room exists in room_excel_file.xlsx
            if exists return True otherwise return False.
        """

        self.load_room_workbook()
        self.room_sheet.dimensions
        self.room_sheet.max_column

        self.room_exists = False

        for r_row in range(1, self.room_sheet.max_row + 1):
            for r_column in range(1, self.room_sheet.max_column + 1):
                if 'No' in self.room_sheet.cell(row=r_row, column=r_column).value:
                    self.room_exists = True

        return self.room_exists

    def check_if_food_exists(self):
        """
            check_if_food_exists method checks whether food exists in food_excel_file.xlsx
            if exists return True otherwise return False.
        """

        self.load_food_workbook()
        self.food_sheet.dimensions
        if self.food_sheet.max_row > 1:
            print('\nAvailable foods\n')
            return True
        else:
            print('\nNot Available Foods\n')
            return False

    def calculate_and_generate_bill(self):
        """
            calculate_and_generate_bill method first check if the customer exists in customer_excel_file.xlsx
            if yes then List or show all the customers and then let the user to select the specific customer
            to calculate the payment and then generate the bill of the customer in pdf format
        """
        
        self.load_customer_workbook()
        self.load_room_workbook()
        self.load_food_workbook()

        self.customer_sheet.dimensions
        self.room_sheet.dimensions
        self.food_sheet.dimensions

        if self.customer_sheet.max_row > 1:
            self.list_customers()
            self.customer_id = input(
                'Enter Customer ID to Generate the bill: ')

            self.customer_room_id = ''
            self.customer_name = ''
            self.customer_address = ''
            self.customer_check_in_date = ''
            self.customer_check_out_date = time.strftime('%Y-%m-%d %H:%m:%S')
            self.customer_food_id = ''
            self.customer_room_no = ''
            self.customer_room_price = ''
            self.customer_food_price = ''

            for c_row in range(1, self.customer_sheet.max_row + 1):
                if self.customer_id in self.customer_sheet['A{index}'.format(index=c_row)].value:
                    self.customer_name = self.customer_sheet['B{index}'.format(
                        index=c_row)].value
                    self.customer_address = self.customer_sheet['C{index}'.format(
                        index=c_row)].value
                    self.customer_check_in_date = self.customer_sheet['D{index}'.format(
                        index=c_row)].value
                    self.customer_room_id = self.customer_sheet['F{index}'.format(
                        index=c_row)].value
                    self.customer_food_id = self.customer_sheet['G{index}'.format(
                        index=c_row)].value
                    break

            for r_row in range(1, self.room_sheet.max_row + 1):

                if self.customer_room_id in self.room_sheet['A{index}'.format(index=r_row)].value:
                    self.customer_room_no = self.room_sheet['B{index}'.format(
                        index=r_row)].value
                    self.customer_room_price = self.room_sheet['C{index}'.format(
                        index=r_row)].value
                    self.room_sheet['D{index}'.format(
                        index=r_row)].value = 'No'
                    break

            for f_row in range(1, self.food_sheet.max_row + 1):
                if self.customer_food_id in self.food_sheet['A{index}'.format(index=f_row)].value:
                    self.customer_food_price = self.food_sheet['C{index}'.format(
                        index=f_row)].value
                    break

            self.pdf_bill_generator(customer_id=self.customer_id, customer_name=self.customer_name, customer_address=self.customer_address, customer_check_in_date=self.customer_check_in_date,
                                    customer_check_out_date=self.customer_check_out_date, customer_room_no=self.customer_room_no, customer_room_rent=self.customer_room_price, customer_food_purchased=self.customer_food_price)

        else:
            print('\nNo Customers available.\n')

    def pdf_bill_generator(self, customer_id, customer_name, customer_address, customer_check_in_date, customer_check_out_date, customer_room_no, customer_room_rent, customer_food_purchased):
        """ pdf_bill_generator actually obrain all the customer, customer food, and customer room's data
        then generate the pdf bill and calcuate some of the date then generate pdf format for the customer. 

        Args:
            customer_id (string): Id of the Customer
            customer_name (String): Name of the Customer
            customer_address (Address): Address of the Customer
            customer_check_in_date (date): The date which customer Entered to Hotel
            customer_check_out_date (date): The current date which Customer 
            customer_room_no (Integer): Room No of Customer
            customer_room_rent (Integer): Rent of the room which the Customer reserved
            customer_food_purchased (Integer): price of the food which customer picked up
        """

        self.pdf = FPDF()
        self.pdf.add_page()
        self.pdf_name = str(customer_name + '_' + customer_id)
        self.total_to_pay = int(
            int(customer_room_rent) + int(customer_food_purchased) + 100)

        self.pdf.set_font("Arial", size=12)
        self.pdf.cell(
            200, 10, txt="***********HOTEL BILL************", ln=1, align="C")
        self.pdf.cell(200, 10, txt="Customer ID: {c_id}".format(
            c_id=customer_id), ln=1, align="C")
        self.pdf.cell(200, 10, txt="Customer Name: {c_name}".format(
            c_name=customer_name), ln=1, align="C")
        self.pdf.cell(200, 10, txt="Customer Address: {c_add}".format(
            c_add=customer_address), ln=1, align="C")
        self.pdf.cell(200, 10, txt="Customer Check In Date: {c_ch_in_date}".format(
            c_ch_in_date=customer_check_in_date), ln=1, align="C")
        self.pdf.cell(200, 10, txt="Customer Check Out Date: {c_ch_out_date}".format(
            c_ch_out_date=customer_check_out_date), ln=1, align="C")
        self.pdf.cell(200, 10, txt="Room No: {c_room_no}".format(
            c_room_no=customer_room_no), ln=1, align="C")
        self.pdf.cell(200, 10, txt="Room Rent: {c_room_rent}".format(
            c_room_rent=customer_room_rent), ln=1, align="C")
        self.pdf.cell(200, 10, txt="Food Purchased Bill: {food_purchased}".format(
            food_purchased=customer_food_purchased), ln=1, align="C")
        self.pdf.cell(200, 10, txt="Food Service Charges: {service_chareges}".format(
            service_chareges='100'), ln=1, align="C")
        self.pdf.cell(200, 10, txt="Total To Pay: {t_t_p}".format(
            t_t_p=self.total_to_pay), ln=1, align="C")
        self.pdf.output("{pdf_name}.pdf".format(pdf_name=self.pdf_name))
